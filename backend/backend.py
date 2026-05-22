import os
from datetime import datetime, timedelta, timezone
import logging
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Depends, Path, Request, Header, File, Form, UploadFile
from fastapi.concurrency import run_in_threadpool
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.middleware import SlowAPIMiddleware
from email_validator import validate_email, EmailNotValidError
from pydantic import BaseModel, EmailStr, SecretStr, Field
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Mapped, Session, mapped_column
import random
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from passlib.context import CryptContext
from langchain_community.utilities import SQLDatabase
from langchain_groq import ChatGroq
import requests
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, ForeignKey, text, or_, inspect as sqlalchemy_inspect
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.types import UserDefinedType
import json
import re
from typing import cast, Any
from typing import List, Optional, Dict
import csv
import io
import hashlib
import secrets
import sqlite3
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime
from sqlalchemy import Column, Integer, String, Text, DateTime
from datetime import datetime
from pathlib import Path as FilePath
from urllib.parse import quote_plus

from extended_models import (
    FavoriteQuery,
    UserSettings,
    QueryHistory,
    Base
)
from sql_system_prompt import SQL_SYSTEM_PROMPT
import re
from typing import Tuple
import threading
import time

# ─────────────────────────────────────────────
# ENV / KEYS
# ─────────────────────────────────────────────
logger = logging.getLogger(__name__)

load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")
if not groq_api_key:
    logger.warning("GROQ_API_KEY not found in environment variables.")
    GROQ_API_KEY = ""
else:
    GROQ_API_KEY = cast(str, groq_api_key)

EMAIL_HOST_USER = os.getenv("EMAIL_HOST_USER")
EMAIL_HOST_PASSWORD = os.getenv("EMAIL_HOST_PASSWORD")
if not EMAIL_HOST_USER or not EMAIL_HOST_PASSWORD:
    logger.warning("Email credentials not found. OTP sending will be disabled.")

USE_OLLAMA = os.getenv("USE_OLLAMA", "true")
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://127.0.0.1:11434").rstrip("/")
OLLAMA_GENERATE_URL = f"{OLLAMA_BASE_URL}/api/generate"
OLLAMA_HEALTH_URL = f"{OLLAMA_BASE_URL}/api/tags"
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "deepseek-coder")
OLLAMA_TIMEOUT_SECONDS = 10
OLLAMA_HEALTH_TIMEOUT_SECONDS = int(os.getenv("OLLAMA_HEALTH_TIMEOUT_SECONDS", "30"))
OLLAMA_MAX_RETRIES = 1

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


class VectorType(UserDefinedType):
    cache_ok = True
    
    def __init__(self, *args, **kwargs):
        self.args = args
        self.kwargs = kwargs

    def get_col_spec(self, **kw):
        if self.args:
            return f"VECTOR({', '.join(str(arg) for arg in self.args)})"
        return "VECTOR"


def register_custom_reflection_types() -> None:
    try:
        from sqlalchemy.dialects.postgresql.base import ischema_names as postgres_ischema_names

        if "vector" not in postgres_ischema_names:
            postgres_ischema_names["vector"] = VectorType
            logger.info("Registered PostgreSQL reflection handler for VECTOR columns.")
    except Exception as exc:
        logger.warning("Unable to register custom reflection types: %s", exc)


register_custom_reflection_types()

# ─────────────────────────────────────────────
# APP-INTERNAL SQLITE DATABASE  (users, sessions, etc.)
# ─────────────────────────────────────────────
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
SQLITE_DB_FILE = os.path.join(BACKEND_DIR, "users.db")
IMPORTED_DB_DIR = os.path.join(BACKEND_DIR, "imported_sources")
os.makedirs(IMPORTED_DB_DIR, exist_ok=True)
engine = create_engine(
    f"sqlite:///{SQLITE_DB_FILE}",
    echo=False,
    pool_pre_ping=True,
    pool_recycle=3600,
    pool_size=10,
    max_overflow=20,
    connect_args={
        "timeout": 30,
        "check_same_thread": False
    },
)


class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, index=True)
    email: Mapped[str] = mapped_column(String, unique=True, index=True, nullable=False)
    firstName: Mapped[str] = mapped_column(String, nullable=False)
    lastName: Mapped[str] = mapped_column(String, nullable=False)
    gender: Mapped[str] = mapped_column(String, nullable=False)
    username: Mapped[str] = mapped_column(String, nullable=False)
    hashed_password: Mapped[str] = mapped_column(String, nullable=False)


class ChatSession(Base):
    __tablename__ = "chat_sessions"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, index=True)
    user_id: Mapped[int] = mapped_column(Integer, ForeignKey('users.id'), nullable=False)
    title: Mapped[str] = mapped_column(String, nullable=False)
    messages: Mapped[str] = mapped_column(Text, nullable=False)


class UserDashboard(Base):
    __tablename__ = "user_dashboards"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    user_id: Mapped[int] = mapped_column(Integer, nullable=False)
    dashboard_id: Mapped[str] = mapped_column(String(100), unique=True, nullable=False)
    name: Mapped[str] = mapped_column(String(200), nullable=False)
    description: Mapped[Optional[str]] = mapped_column(Text)
    charts_data: Mapped[str] = mapped_column(Text, nullable=False)
    created_at: Mapped[Optional[datetime]] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[Optional[datetime]] = mapped_column(DateTime, default=datetime.utcnow)


class AuthSession(Base):
    __tablename__ = "auth_sessions"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, index=True)
    user_id: Mapped[int] = mapped_column(Integer, ForeignKey('users.id'), nullable=False, index=True)
    token_hash: Mapped[str] = mapped_column(String(64), unique=True, index=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, nullable=False)
    expires_at: Mapped[datetime] = mapped_column(DateTime, nullable=False)


class DatabaseSession(Base):
    __tablename__ = "database_sessions"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, index=True)
    user_id: Mapped[int] = mapped_column(Integer, ForeignKey('users.id'), nullable=False, index=True)
    token_hash: Mapped[str] = mapped_column(String(64), unique=True, index=True, nullable=False)
    db_uri: Mapped[str] = mapped_column(Text, nullable=False)
    db_name: Mapped[str] = mapped_column(String, nullable=False)
    db_type: Mapped[str] = mapped_column(String(32), nullable=False, default="mysql")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, nullable=False)
    expires_at: Mapped[datetime] = mapped_column(DateTime, nullable=False)


class OtpCode(Base):
    __tablename__ = "otp_codes"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, index=True)
    email: Mapped[str] = mapped_column(String, unique=True, index=True, nullable=False)
    otp_hash: Mapped[str] = mapped_column(String, nullable=False)
    attempts: Mapped[int] = mapped_column(Integer, default=0, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, nullable=False)
    expires_at: Mapped[datetime] = mapped_column(DateTime, nullable=False)


def ensure_internal_schema() -> None:
    Base.metadata.create_all(engine)
    with engine.begin() as connection:
        inspector = sqlalchemy_inspect(connection)
        if inspector.has_table("database_sessions"):
            columns = {column["name"] for column in inspector.get_columns("database_sessions")}
            if "db_type" not in columns:
                connection.execute(
                    text("ALTER TABLE database_sessions ADD COLUMN db_type VARCHAR(32) NOT NULL DEFAULT 'mysql'")
                )


ensure_internal_schema()
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# ─────────────────────────────────────────────
# PER-USER DATABASE SESSION STORE
# Replaces the dangerous global app.state.db_uri
# Each user gets their own db_uri keyed by a session token
# that is issued at /api/connect and must be sent as
# X-DB-Session header on every subsequent request.
# ─────────────────────────────────────────────

class DbSessionStore:
    """Database-backed, per-user database connection store."""

    def __init__(self, ttl_seconds: int = 3600):
        self._ttl = ttl_seconds

    def _cleanup_expired(self, db: Session) -> None:
        db.query(DatabaseSession).filter(DatabaseSession.expires_at <= datetime.utcnow()).delete()
        db.flush()

    def create(self, user_id: int, db_uri: str, db_name: str, db_type: str, db: Session) -> str:
        """Store a new connection and return an opaque session token."""
        token = secrets.token_urlsafe(32)
        self._cleanup_expired(db)
        existing_sessions = db.query(DatabaseSession).filter(DatabaseSession.user_id == user_id).all()
        for existing_session in existing_sessions:
            cleanup_imported_snapshot(existing_session.db_uri, existing_session.db_type)
            db.delete(existing_session)
        db.add(
            DatabaseSession(
                user_id=user_id,
                token_hash=hash_token(token),
                db_uri=db_uri,
                db_name=db_name,
                db_type=normalize_source_type(db_type),
                expires_at=datetime.utcnow() + timedelta(seconds=self._ttl),
            )
        )
        db.commit()
        return token

    def get(self, token: str, db: Session, user_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        """Return the session dict or None if missing / expired."""
        query = db.query(DatabaseSession).filter(
            DatabaseSession.token_hash == hash_token(token),
            DatabaseSession.expires_at > datetime.utcnow(),
        )
        if user_id is not None:
            query = query.filter(DatabaseSession.user_id == user_id)
        session = query.first()
        if session is None:
            return None
        return {
            "db_uri": session.db_uri,
            "db_name": session.db_name,
            "db_type": session.db_type,
            "session_id": session.id,
            "user_id": session.user_id,
        }

    def delete(self, token: str, db: Session, user_id: Optional[int] = None) -> None:
        query = db.query(DatabaseSession).filter(DatabaseSession.token_hash == hash_token(token))
        if user_id is not None:
            query = query.filter(DatabaseSession.user_id == user_id)
        for session in query.all():
            cleanup_imported_snapshot(session.db_uri, session.db_type)
            db.delete(session)
        db.commit()


db_session_store = DbSessionStore(ttl_seconds=int(os.getenv("DB_SESSION_TTL", "3600")))


# ─────────────────────────────────────────────
# QUERY RESULT ROW LIMIT
# ─────────────────────────────────────────────
DEFAULT_MAX_RESULT_ROWS = 10000
MAX_RESULT_ROWS = max(
    1,
    min(int(os.getenv("MAX_RESULT_ROWS", str(DEFAULT_MAX_RESULT_ROWS))), DEFAULT_MAX_RESULT_ROWS),
)


# ─────────────────────────────────────────────
# QUERY CACHE  (scoped per db_name, not global)
# ─────────────────────────────────────────────
class QueryCache:
    """In-process TTL cache for SELECT results."""

    def __init__(self, ttl_seconds: int = 300, max_entries: int = 200):
        self.cache: Dict[str, Any] = {}
        self.timestamps: Dict[str, float] = {}
        self.ttl = ttl_seconds
        self.max_entries = max_entries
        self._lock = threading.Lock()

    def _key(self, sql: str, db_name: str) -> str:
        return hashlib.md5(f"{sql}:{db_name}".encode()).hexdigest()

    def get(self, sql: str, db_name: str) -> Optional[dict]:
        key = self._key(sql, db_name)
        with self._lock:
            if key not in self.cache:
                return None
            if time.time() - self.timestamps[key] > self.ttl:
                del self.cache[key]
                del self.timestamps[key]
                return None
            return self.cache[key]

    def set(self, sql: str, db_name: str, value: dict) -> None:
        key = self._key(sql, db_name)
        with self._lock:
            # Evict oldest entry when over capacity
            if len(self.cache) >= self.max_entries and key not in self.cache:
                oldest = min(self.timestamps, key=lambda k: self.timestamps[k])
                del self.cache[oldest]
                del self.timestamps[oldest]
            self.cache[key] = value
            self.timestamps[key] = time.time()

    def clear(self) -> None:
        with self._lock:
            self.cache.clear()
            self.timestamps.clear()


query_cache = QueryCache(ttl_seconds=300, max_entries=200)

# ─────────────────────────────────────────────
# FASTAPI APP
# ─────────────────────────────────────────────
app = FastAPI()

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter

DEFAULT_ALLOWED_ORIGINS = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:8081",
    "http://127.0.0.1:8081",
    "http://localhost:8082",
    "http://127.0.0.1:8082",
]

allowed_origins_env = os.getenv("ALLOWED_ORIGINS")
ALLOWED_ORIGINS = (
    [origin.strip() for origin in allowed_origins_env.split(",") if origin.strip()]
    if allowed_origins_env
    else DEFAULT_ALLOWED_ORIGINS
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(SlowAPIMiddleware)

# ─────────────────────────────────────────────
# OTP MANAGER
# ─────────────────────────────────────────────
class OtpManager:
    """Database-backed OTP manager."""

    def _cleanup_expired(self, db: Session) -> None:
        db.query(OtpCode).filter(OtpCode.expires_at <= datetime.utcnow()).delete()
        db.flush()

    def store(self, email: str, otp: str, db: Session, expiry_minutes: int = 5) -> None:
        self._cleanup_expired(db)
        existing = db.query(OtpCode).filter(OtpCode.email == email).first()
        if existing:
            db.delete(existing)
            db.flush()
        db.add(
            OtpCode(
                email=email,
                otp_hash=get_password_hash(otp),
                attempts=0,
                expires_at=datetime.utcnow() + timedelta(minutes=expiry_minutes),
            )
        )
        db.commit()

    def verify(self, email: str, otp: str, db: Session) -> Tuple[bool, str]:
        self._cleanup_expired(db)
        otp_record = db.query(OtpCode).filter(OtpCode.email == email).first()
        if otp_record is None:
            return False, "OTP not requested or already used"
        if otp_record.expires_at <= datetime.utcnow():
            db.delete(otp_record)
            db.commit()
            return False, "OTP has expired"
        if otp_record.attempts >= 5:
            db.delete(otp_record)
            db.commit()
            return False, "Too many failed attempts. Request a new OTP."
        if not verify_password(otp, otp_record.otp_hash):
            otp_record.attempts += 1
            db.commit()
            return False, "Invalid OTP"
        db.delete(otp_record)
        db.commit()
        return True, "OTP verified"


otp_manager = OtpManager()

# ─────────────────────────────────────────────
# PYDANTIC MODELS
# ─────────────────────────────────────────────
class DBConfig(BaseModel):
    type: str = "mysql"
    host: str = ""
    port: Optional[int] = None
    user: str = ""
    password: str = ""
    database: str = ""
    path: str = ""


class ServerConnectionConfig(BaseModel):
    type: str = "mysql"
    host: str = ""
    port: Optional[int] = None
    user: str = ""
    password: str = ""


class CreateDatabaseRequest(BaseModel):
    type: str = "mysql"
    host: str = ""
    port: Optional[int] = None
    user: str = ""
    password: str = ""
    database_name: str


class ChatRequest(BaseModel):
    question: str
    chat_history: list = Field(default_factory=list)


class UserCreate(BaseModel):
    firstName: str
    lastName: str
    email: EmailStr
    password: str
    otp: str
    gender: str
    username: str


class UserLogin(BaseModel):
    identifier: str
    password: str


class OtpRequest(BaseModel):
    email: EmailStr


class FavoriteQueryCreate(BaseModel):
    user_id: Optional[int] = None
    question: str
    sql_query: str
    tags: Optional[str] = None
    description: Optional[str] = None


class ChatSessionCreateRequest(BaseModel):
    title: str = "Untitled Chat"
    messages: List[Dict[str, Any]] = Field(default_factory=list)
    user_id: Optional[int] = None


class ChatSessionUpdateRequest(BaseModel):
    title: Optional[str] = None
    messages: Optional[List[Dict[str, Any]]] = None
    user_id: Optional[int] = None


class QueryHistoryTrackRequest(BaseModel):
    user_id: Optional[int] = None
    session_id: Optional[int] = None
    question: str
    sql_query: str
    success: bool = True
    execution_time_ms: Optional[int] = None
    row_count: Optional[int] = None


class UserSettingsUpdate(BaseModel):
    theme: Optional[str] = None
    language: Optional[str] = None
    results_per_page: Optional[int] = None
    auto_save_sessions: Optional[bool] = None
    sql_syntax_highlighting: Optional[bool] = None
    notification_preferences: Optional[dict] = None


class ExportRequest(BaseModel):
    data: List[List]
    columns: List[str]
    format: str


class ChartConfig(BaseModel):
    xKey: Optional[str] = "name"
    yKey: Optional[str] = "value"
    dataKey: Optional[str] = "value"
    nameKey: Optional[str] = "name"
    colors: Optional[List[str]] = None


class ChartData(BaseModel):
    id: str
    title: str
    type: str
    size: Optional[str] = "medium"
    data: List[Dict[str, Any]]
    config: ChartConfig
    notes: Optional[str] = None
    createdAt: str
    updatedAt: Optional[str] = None


class Dashboard(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    charts: List[ChartData]
    createdAt: str
    updatedAt: str


class DashboardCreate(BaseModel):
    user_id: Optional[int] = None
    dashboard_id: str
    name: str
    description: Optional[str] = ""
    charts: List[Dict[str, Any]]


class DashboardUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    charts: Optional[List[Dict[str, Any]]] = None


class DashboardMigrate(BaseModel):
    user_id: Optional[int] = None
    dashboards_data: List[Dict[str, Any]]


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


AUTH_SESSION_TTL_SECONDS = int(os.getenv("AUTH_SESSION_TTL", "604800"))
# SHOW, DESCRIBE, DESC removed — system prompt instructs LLM to use
# INFORMATION_SCHEMA instead; allowing them here contradicts that rule
# and causes the LLM to emit shortcuts instead of proper SELECT queries.
READ_ONLY_SQL_PREFIXES = ("SELECT", "WITH", "EXPLAIN", "PRAGMA")

# Per-dialect injection rules — added to every LLM prompt so the model
# always knows exactly which SQL syntax to produce.
DIALECT_RULES = {
    "mysql": (
        "You are generating MySQL 8.x SQL.\n"
        "Rules:\n"
        "- Use LIMIT N (not FETCH FIRST)\n"
        "- Use INFORMATION_SCHEMA.TABLES / INFORMATION_SCHEMA.COLUMNS for metadata\n"
        "- Never use SHOW TABLES or DESCRIBE\n"
        "- Use DATE(), CURDATE(), NOW(), DATE_SUB() for dates\n"
        "- Use CONCAT() for string concatenation\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "mariadb": (
        "You are generating MariaDB 10.x SQL.\n"
        "Rules:\n"
        "- Use LIMIT N (not FETCH FIRST)\n"
        "- Use INFORMATION_SCHEMA for metadata\n"
        "- Never use SHOW TABLES or DESCRIBE\n"
        "- Use DATE(), CURDATE(), NOW() for dates\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "postgresql": (
        "You are generating PostgreSQL 15+ SQL.\n"
        "Rules:\n"
        "- Use LIMIT N / OFFSET for pagination\n"
        "- Use information_schema.tables / information_schema.columns for metadata\n"
        "- Use ILIKE for case-insensitive LIKE\n"
        "- Use NOW() - INTERVAL 'N days' for date arithmetic\n"
        "- Use col::TEXT for casting\n"
        "- Use || for string concatenation\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "oracle": (
        "You are generating Oracle 12c+ SQL.\n"
        "Rules:\n"
        "- Use FETCH FIRST N ROWS ONLY (not LIMIT)\n"
        "- Use user_tables / user_tab_columns for metadata\n"
        "- Use SYSDATE for current date, TO_DATE() for literals\n"
        "- Use || for string concatenation\n"
        "- Use NVL() for null handling\n"
        "- Use DUAL for scalar expressions: SELECT ... FROM dual\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "mssql": (
        "You are generating Microsoft SQL Server T-SQL.\n"
        "Rules:\n"
        "- Use TOP(N) or FETCH FIRST N ROWS ONLY for row limits\n"
        "- Use INFORMATION_SCHEMA for metadata\n"
        "- Use GETDATE() for current datetime, DATEADD() for arithmetic\n"
        "- Use ISNULL() or COALESCE() for null handling\n"
        "- Use N'value' prefix for Unicode string literals\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "db2": (
        "You are generating IBM Db2 SQL.\n"
        "Rules:\n"
        "- Use FETCH FIRST N ROWS ONLY (not LIMIT)\n"
        "- Use SYSCAT.TABLES / SYSCAT.COLUMNS for metadata\n"
        "- Use CURRENT DATE / CURRENT TIMESTAMP (no parentheses)\n"
        "- Use || for string concatenation\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
    "sqlite": (
        "You are generating SQLite 3 SQL.\n"
        "Rules:\n"
        "- Use LIMIT N / OFFSET for pagination\n"
        "- Use sqlite_master WHERE type='table' for table metadata\n"
        "- Use PRAGMA table_info(table_name) for column metadata\n"
        "- Use DATE('now', '-N days') for date arithmetic\n"
        "- Use strftime('%Y', col) to extract date parts\n"
        "- Always include GROUP BY for every non-aggregate SELECT column"
    ),
}

DEFAULT_DIALECT_RULES = (
    "You are generating standard SQL.\n"
    "Rules:\n"
    "- Use LIMIT N for row limits\n"
    "- Use INFORMATION_SCHEMA for metadata\n"
    "- Always include GROUP BY for every non-aggregate SELECT column"
)

SQL_SOURCE_TYPES = {"mysql", "postgresql", "mariadb", "oracle", "sqlserver", "db2", "sqlite", "csv", "excel"}
FILE_SOURCE_TYPES = {"csv", "excel"}
METADATA_ONLY_SOURCE_TYPES = {"mongodb", "redis"}
CHAT_ENABLED_SOURCE_TYPES = SQL_SOURCE_TYPES


def hash_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def normalize_source_type(source_type: Optional[str]) -> str:
    normalized = (source_type or "mysql").strip().lower()
    aliases = {
        "postgres": "postgresql",
        "postgresql": "postgresql",
        "mssql": "sqlserver",
        "sql_server": "sqlserver",
        "mariadb": "mariadb",
        "db2": "db2",
        "excel/csv": "csv",
    }
    return aliases.get(normalized, normalized)


def source_supports_query(source_type: Optional[str]) -> bool:
    return normalize_source_type(source_type) in CHAT_ENABLED_SOURCE_TYPES


def quote_sql_value(value: str) -> str:
    return quote_plus(value or "")


def build_sqlalchemy_uri(
    source_type: str,
    host: str = "",
    port: Optional[int] = None,
    user: str = "",
    password: str = "",
    database: str = "",
    path: str = "",
) -> str:
    normalized = normalize_source_type(source_type)
    if normalized == "sqlite":
        sqlite_path = os.path.abspath(path or database)
        if not sqlite_path:
            raise ValueError("A SQLite file path is required.")
        return f"sqlite:///{sqlite_path.replace(os.sep, '/')}"

    if not host:
        raise ValueError("Host is required.")
    if port is None:
        raise ValueError("Port is required.")

    user_part = quote_sql_value(user)
    password_part = quote_sql_value(password)
    database_part = quote_sql_value(database)

    if normalized in {"mysql", "mariadb"}:
        return (
            f"mysql+mysqlconnector://{user_part}:{password_part}"
            f"@{host}:{port}/{database_part}"
        )
    if normalized == "postgresql":
        return (
            f"postgresql+psycopg2://{user_part}:{password_part}"
            f"@{host}:{port}/{database_part}"
        )
    if normalized == "oracle":
        return (
            f"oracle+oracledb://{user_part}:{password_part}"
            f"@{host}:{port}/?service_name={database_part}"
        )
    if normalized == "sqlserver":
        odbc_connect = (
            "DRIVER={ODBC Driver 18 for SQL Server};"
            f"SERVER={host},{port};"
            f"DATABASE={database};"
            f"UID={user};"
            f"PWD={password};"
            "TrustServerCertificate=yes;"
            "Encrypt=no"
        )
        return f"mssql+pyodbc:///?odbc_connect={quote_plus(odbc_connect)}"
    if normalized == "db2":
        return (
            f"db2+ibm_db://{user_part}:{password_part}"
            f"@{host}:{port}/{database_part}"
        )
    raise ValueError(f"Unsupported SQL source type: {source_type}")


def build_mongodb_uri(
    host: str,
    port: Optional[int],
    user: str = "",
    password: str = "",
    database: str = "admin",
) -> str:
    if not host:
        raise ValueError("Host is required.")
    if port is None:
        raise ValueError("Port is required.")
    auth = ""
    if user:
        auth = quote_sql_value(user)
        if password:
            auth += f":{quote_sql_value(password)}"
        auth += "@"
    elif password:
        auth = f":{quote_sql_value(password)}@"
    return f"mongodb://{auth}{host}:{port}/{quote_sql_value(database or 'admin')}"


def build_redis_uri(
    host: str,
    port: Optional[int],
    user: str = "",
    password: str = "",
    database: str = "0",
) -> str:
    if not host:
        raise ValueError("Host is required.")
    if port is None:
        raise ValueError("Port is required.")
    auth = ""
    if user:
        auth = quote_sql_value(user)
        if password:
            auth += f":{quote_sql_value(password)}"
        auth += "@"
    elif password:
        auth = f":{quote_sql_value(password)}@"
    return f"redis://{auth}{host}:{port}/{database or '0'}"


def slugify_identifier(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z_]+", "_", value.strip().lower()).strip("_")
    if not cleaned:
        cleaned = fallback
    if cleaned[0].isdigit():
        cleaned = f"t_{cleaned}"
    return cleaned[:63]


def unique_identifiers(raw_values: List[str], prefix: str) -> List[str]:
    seen: Dict[str, int] = {}
    normalized: List[str] = []
    for index, raw_value in enumerate(raw_values, start=1):
        base = slugify_identifier(raw_value or f"{prefix}_{index}", f"{prefix}_{index}")
        count = seen.get(base, 0)
        seen[base] = count + 1
        normalized.append(base if count == 0 else f"{base}_{count + 1}")
    return normalized


def sqlite_identifier(name: str) -> str:
    return f'"{name.replace(chr(34), chr(34) * 2)}"'


def create_imported_sqlite_path(label: str, user_id: int) -> str:
    safe_label = slugify_identifier(FilePath(label).stem, "uploaded_source")
    filename = f"{safe_label}_{user_id}_{int(time.time())}_{secrets.token_hex(4)}.sqlite"
    return os.path.join(IMPORTED_DB_DIR, filename)


def cleanup_imported_snapshot(db_uri: str, db_type: Optional[str]) -> None:
    if normalize_source_type(db_type) not in FILE_SOURCE_TYPES:
        return
    if not db_uri.startswith("sqlite:///"):
        return
    db_path = db_uri.replace("sqlite:///", "", 1)
    try:
        resolved = os.path.abspath(db_path)
        if os.path.commonpath([resolved, os.path.abspath(IMPORTED_DB_DIR)]) == os.path.abspath(IMPORTED_DB_DIR):
            if os.path.exists(resolved):
                os.remove(resolved)
    except Exception:
        pass


def create_sqlite_snapshot_from_csv(file_name: str, file_bytes: bytes, user_id: int) -> str:
    csv_text = file_bytes.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(csv_text)))
    if not reader:
        raise ValueError("The CSV file is empty.")

    headers = unique_identifiers([str(value) for value in reader[0]], "column")
    rows = reader[1:]
    table_name = slugify_identifier(FilePath(file_name).stem, "sheet1")
    sqlite_path = create_imported_sqlite_path(file_name, user_id)

    connection = sqlite3.connect(sqlite_path)
    try:
        columns_sql = ", ".join(f"{sqlite_identifier(column)} TEXT" for column in headers)
        connection.execute(f"CREATE TABLE {sqlite_identifier(table_name)} ({columns_sql})")
        placeholders = ", ".join("?" for _ in headers)
        normalized_rows = []
        for row in rows:
            padded = list(row[: len(headers)]) + [""] * max(0, len(headers) - len(row))
            normalized_rows.append([("" if cell is None else str(cell)) for cell in padded[: len(headers)]])
        if normalized_rows:
            connection.executemany(
                f"INSERT INTO {sqlite_identifier(table_name)} VALUES ({placeholders})",
                normalized_rows,
            )
        connection.commit()
    finally:
        connection.close()

    return f"sqlite:///{sqlite_path.replace(os.sep, '/')}"


def create_sqlite_snapshot_from_excel(file_name: str, file_bytes: bytes, user_id: int) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel import requires the openpyxl package.") from exc

    sqlite_path = create_imported_sqlite_path(file_name, user_id)
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    connection = sqlite3.connect(sqlite_path)

    try:
        created_tables = 0
        for sheet in workbook.worksheets:
            sheet_rows = list(sheet.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            headers = unique_identifiers([str(value or "") for value in sheet_rows[0]], "column")
            table_name = slugify_identifier(sheet.title, f"sheet_{created_tables + 1}")
            columns_sql = ", ".join(f"{sqlite_identifier(column)} TEXT" for column in headers)
            connection.execute(f"CREATE TABLE {sqlite_identifier(table_name)} ({columns_sql})")
            placeholders = ", ".join("?" for _ in headers)
            normalized_rows = []
            for row in sheet_rows[1:]:
                padded = list(row[: len(headers)]) + [""] * max(0, len(headers) - len(row))
                normalized_rows.append([("" if cell is None else str(cell)) for cell in padded[: len(headers)]])
            if normalized_rows:
                connection.executemany(
                    f"INSERT INTO {sqlite_identifier(table_name)} VALUES ({placeholders})",
                    normalized_rows,
                )
            created_tables += 1
        if created_tables == 0:
            raise ValueError("The Excel workbook does not contain any populated sheets.")
        connection.commit()
    finally:
        connection.close()

    return f"sqlite:///{sqlite_path.replace(os.sep, '/')}"


def create_sql_snapshot_for_uploaded_file(source_type: str, upload: UploadFile, user_id: int) -> str:
    file_bytes = upload.file.read()
    if not file_bytes:
        raise ValueError("The uploaded file is empty.")
    if normalize_source_type(source_type) == "csv":
        return create_sqlite_snapshot_from_csv(upload.filename or "uploaded.csv", file_bytes, user_id)
    if normalize_source_type(source_type) == "excel":
        return create_sqlite_snapshot_from_excel(upload.filename or "uploaded.xlsx", file_bytes, user_id)
    raise ValueError(f"Unsupported uploaded source type: {source_type}")


def create_auth_token(user_id: int, db: Session) -> str:
    token = secrets.token_urlsafe(32)
    db.query(AuthSession).filter(AuthSession.expires_at <= datetime.utcnow()).delete()
    db.add(
        AuthSession(
            user_id=user_id,
            token_hash=hash_token(token),
            expires_at=datetime.utcnow() + timedelta(seconds=AUTH_SESSION_TTL_SECONDS),
        )
    )
    db.commit()
    return token


def extract_bearer_token(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header is required")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Invalid authorization header")
    return token


def get_current_user(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
) -> User:
    token = extract_bearer_token(authorization)
    token_hash = hash_token(token)
    auth_session = db.query(AuthSession).filter(AuthSession.token_hash == token_hash).first()
    if auth_session is None:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    if auth_session.expires_at <= datetime.utcnow():
        db.delete(auth_session)
        db.commit()
        raise HTTPException(status_code=401, detail="Session expired. Please log in again.")
    user = db.query(User).filter(User.id == auth_session.user_id).first()
    if user is None:
        db.delete(auth_session)
        db.commit()
        raise HTTPException(status_code=401, detail="User not found for this session")
    return user


def revoke_auth_token(token: str, db: Session) -> None:
    db.query(AuthSession).filter(AuthSession.token_hash == hash_token(token)).delete()
    db.commit()


def assert_user_access(requested_user_id: Optional[int], current_user: User) -> None:
    if requested_user_id is not None and requested_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="You are not authorized to access another user's data")


def get_db_session(
    x_db_session: Optional[str] = Header(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """FastAPI dependency – resolves the per-user DB session from the header."""
    if not x_db_session:
        raise HTTPException(status_code=400, detail="X-DB-Session header is required. Connect to a database first.")
    session = db_session_store.get(x_db_session, db, current_user.id)
    if session is None:
        raise HTTPException(status_code=401, detail="Database session expired or not found. Please reconnect.")
    return session


class EngineCache:
    def __init__(self):
        self._engines: Dict[str, Any] = {}
        self._lock = threading.Lock()

    def get(self, db_uri: str):
        with self._lock:
            engine = self._engines.get(db_uri)
            if engine is None:
                engine_kwargs: Dict[str, Any] = {
                    "pool_pre_ping": True,
                    "pool_recycle": 3600,
                }
                if db_uri.startswith("sqlite:///"):
                    engine_kwargs["connect_args"] = {"check_same_thread": False}
                else:
                    engine_kwargs["pool_size"] = 10
                    engine_kwargs["max_overflow"] = 20
                engine = create_engine(db_uri, **engine_kwargs)
                self._engines[db_uri] = engine
            return engine

    def dispose_all(self) -> None:
        with self._lock:
            for engine in self._engines.values():
                engine.dispose()
            self._engines.clear()


engine_cache = EngineCache()


def validate_source_connection(db_uri: str, source_type: str, db_name: str = "") -> None:
    normalized = normalize_source_type(source_type)
    if normalized in SQL_SOURCE_TYPES:
        engine = create_engine(
            db_uri,
            connect_args={"check_same_thread": False} if db_uri.startswith("sqlite:///") else {},
            pool_pre_ping=True,
        )
        try:
            with engine.connect() as connection:
                connection.execute(text("SELECT 1"))
        finally:
            engine.dispose()
        return

    if normalized == "mongodb":
        try:
            from pymongo import MongoClient
        except ImportError as exc:
            raise ValueError("MongoDB support requires the pymongo package.") from exc

        client = MongoClient(db_uri, serverSelectionTimeoutMS=5000)
        try:
            client.admin.command("ping")
            if db_name:
                client[db_name].list_collection_names()
        finally:
            client.close()
        return

    if normalized == "redis":
        try:
            import redis
        except ImportError as exc:
            raise ValueError("Redis support requires the redis package.") from exc

        client = redis.Redis.from_url(db_uri, socket_connect_timeout=5, socket_timeout=5, decode_responses=True)
        client.ping()
        return

    raise ValueError(f"Unsupported source type: {source_type}")


def list_available_databases_for_source(config: ServerConnectionConfig) -> List[str]:
    source_type = normalize_source_type(config.type)
    if source_type in {"mysql", "mariadb"}:
        engine = create_engine(
            build_sqlalchemy_uri(source_type, config.host, config.port, config.user, config.password, "mysql")
        )
        try:
            with engine.connect() as connection:
                result = connection.execute(text("SHOW DATABASES"))
                databases = [str(row[0]) for row in result.fetchall()]
        finally:
            engine.dispose()
        system_dbs = {"information_schema", "mysql", "performance_schema", "sys"}
        return [name for name in databases if name not in system_dbs]

    if source_type == "postgresql":
        engine = create_engine(
            build_sqlalchemy_uri(source_type, config.host, config.port, config.user, config.password, "postgres")
        )
        try:
            with engine.connect() as connection:
                result = connection.execute(
                    text(
                        "SELECT datname FROM pg_database "
                        "WHERE datistemplate = false ORDER BY datname"
                    )
                )
                return [str(row[0]) for row in result.fetchall()]
        finally:
            engine.dispose()

    if source_type == "sqlserver":
        engine = create_engine(
            build_sqlalchemy_uri(source_type, config.host, config.port, config.user, config.password, "master")
        )
        try:
            with engine.connect() as connection:
                result = connection.execute(
                    text("SELECT name FROM sys.databases WHERE database_id > 4 ORDER BY name")
                )
                return [str(row[0]) for row in result.fetchall()]
        finally:
            engine.dispose()

    if source_type == "mongodb":
        try:
            from pymongo import MongoClient
        except ImportError as exc:
            raise ValueError("MongoDB support requires the pymongo package.") from exc

        client = MongoClient(
            build_mongodb_uri(config.host, config.port, config.user, config.password, "admin"),
            serverSelectionTimeoutMS=5000,
        )
        try:
            names = client.list_database_names()
            return [name for name in names if name not in {"admin", "config", "local"}]
        finally:
            client.close()

    raise ValueError(f"Listing databases is not available for {config.type}.")


def create_database_for_source(request: CreateDatabaseRequest) -> None:
    source_type = normalize_source_type(request.type)
    if not re.match(r"^[a-zA-Z0-9_]+$", request.database_name):
        raise ValueError("Database name can only contain letters, numbers, and underscores.")

    if source_type in {"mysql", "mariadb"}:
        engine = create_engine(
            build_sqlalchemy_uri(source_type, request.host, request.port, request.user, request.password, "mysql")
        )
        try:
            with engine.begin() as connection:
                connection.execute(text(f"CREATE DATABASE IF NOT EXISTS `{request.database_name}`"))
        finally:
            engine.dispose()
        return

    if source_type == "postgresql":
        engine = create_engine(
            build_sqlalchemy_uri(source_type, request.host, request.port, request.user, request.password, "postgres"),
            isolation_level="AUTOCOMMIT",
        )
        try:
            with engine.connect() as connection:
                connection.execute(text(f'CREATE DATABASE "{request.database_name}"'))
        finally:
            engine.dispose()
        return

    raise ValueError(f"Creating databases is not available for {request.type}.")


def build_connection_artifacts(config: DBConfig) -> Tuple[str, str, str]:
    source_type = normalize_source_type(config.type)

    if source_type in SQL_SOURCE_TYPES:
        db_uri = build_sqlalchemy_uri(
            source_type,
            config.host,
            config.port,
            config.user,
            config.password,
            config.database,
            config.path,
        )
        db_name = config.database or FilePath(config.path or "sqlite").name
        return source_type, db_uri, db_name

    if source_type == "mongodb":
        return (
            source_type,
            build_mongodb_uri(config.host, config.port, config.user, config.password, config.database or "admin"),
            config.database or "admin",
        )

    if source_type == "redis":
        return (
            source_type,
            build_redis_uri(config.host, config.port, config.user, config.password, config.database or "0"),
            config.database or "0",
        )

    raise ValueError(f"Unsupported source type: {config.type}")


def get_sql_tables_info(db_uri: str) -> List[Dict[str, Any]]:
    engine = engine_cache.get(db_uri)
    inspector = sqlalchemy_inspect(engine)
    table_names = inspector.get_table_names()
    preparer = engine.dialect.identifier_preparer

    tables: List[Dict[str, Any]] = []
    with engine.connect() as connection:
        for table_name in table_names:
            quoted_table = preparer.quote_identifier(table_name)
            row_count = 0
            try:
                row_count = int(connection.execute(text(f"SELECT COUNT(*) FROM {quoted_table}")).scalar() or 0)
            except Exception as table_error:
                print(f"Error counting rows for table {table_name}: {table_error}")
            tables.append({"name": table_name, "rowCount": row_count, "lastUpdated": "unknown"})
    return tables


def get_sql_table_schema(db_uri: str, table_name: str) -> List[Dict[str, Any]]:
    engine = engine_cache.get(db_uri)
    inspector = sqlalchemy_inspect(engine)
    columns = inspector.get_columns(table_name)
    return [
        {
            "name": column["name"],
            "type": str(column["type"]),
            "nullable": bool(column.get("nullable", True)),
            "default": column.get("default"),
            "autoincrement": bool(column.get("autoincrement")),
        }
        for column in columns
    ]


def get_mongodb_tables_info(db_uri: str, db_name: str) -> List[Dict[str, Any]]:
    try:
        from pymongo import MongoClient
    except ImportError as exc:
        raise ValueError("MongoDB support requires the pymongo package.") from exc

    client = MongoClient(db_uri, serverSelectionTimeoutMS=5000)
    try:
        database = client[db_name]
        tables = []
        for collection_name in database.list_collection_names():
            count = database[collection_name].estimated_document_count()
            tables.append({"name": collection_name, "rowCount": int(count), "lastUpdated": "live"})
        return tables
    finally:
        client.close()


def get_mongodb_table_schema(db_uri: str, db_name: str, collection_name: str) -> List[Dict[str, Any]]:
    try:
        from pymongo import MongoClient
    except ImportError as exc:
        raise ValueError("MongoDB support requires the pymongo package.") from exc

    client = MongoClient(db_uri, serverSelectionTimeoutMS=5000)
    try:
        sample = client[db_name][collection_name].find_one() or {}
        return [
            {
                "name": field_name,
                "type": type(field_value).__name__,
                "nullable": True,
                "default": None,
                "autoincrement": False,
            }
            for field_name, field_value in sample.items()
        ]
    finally:
        client.close()


def get_redis_tables_info(db_uri: str) -> List[Dict[str, Any]]:
    try:
        import redis
    except ImportError as exc:
        raise ValueError("Redis support requires the redis package.") from exc

    client = redis.Redis.from_url(db_uri, socket_connect_timeout=5, socket_timeout=5, decode_responses=True)
    dbsize_result = cast(Any, client.dbsize())
    return [{"name": "redis_keys", "rowCount": int(dbsize_result), "lastUpdated": "live"}]


def get_redis_table_schema() -> List[Dict[str, Any]]:
    return [
        {"name": "key", "type": "string", "nullable": False, "default": None, "autoincrement": False},
        {"name": "type", "type": "string", "nullable": True, "default": None, "autoincrement": False},
        {"name": "ttl", "type": "integer", "nullable": True, "default": None, "autoincrement": False},
        {"name": "value_preview", "type": "string", "nullable": True, "default": None, "autoincrement": False},
    ]


def generate_otp():
    return str(random.randint(100000, 999999))


def send_otp_email(recipient_email: str, otp: str):
    if not EMAIL_HOST_USER or not EMAIL_HOST_PASSWORD:
        logger.warning("Email credentials missing; OTP was not sent to %s", recipient_email)
        raise HTTPException(
            status_code=503,
            detail="Email OTP is not configured. Set EMAIL_HOST_USER and EMAIL_HOST_PASSWORD.",
        )

    message = MIMEMultipart("alternative")
    message["Subject"] = "Your Verification Code"
    message["From"] = EMAIL_HOST_USER
    message["To"] = recipient_email

    html = f"""
    <html>
    <body>
        <div style="font-family: Arial, sans-serif; text-align: center; color: #333;">
            <h2>Welcome to Query Genie!</h2>
            <p>Your one-time verification code is:</p>
            <p style="font-size: 24px; font-weight: bold; letter-spacing: 2px; color: #007BFF;">{otp}</p>
            <p>This code will expire in 5 minutes.</p>
        </div>
    </body>
    </html>
    """
    message.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
            server.sendmail(EMAIL_HOST_USER, recipient_email, message.as_string())
    except smtplib.SMTPAuthenticationError as e:
        logger.warning("SMTP authentication failed for %s: %s", EMAIL_HOST_USER, e)
        raise HTTPException(
            status_code=503,
            detail="SMTP authentication failed. For Gmail, use a Google app password for EMAIL_HOST_PASSWORD.",
        )
    except Exception as e:
        logger.exception("Failed to send OTP email to %s: %s", recipient_email, e)
        raise HTTPException(status_code=502, detail="Failed to send OTP email.")


DANGEROUS_KEYWORDS = [
    "DROP",
    "TRUNCATE",
    "DELETE",
    "ALTER",
    "UPDATE",
    "INSERT",
    "CREATE",
    "REPLACE",
    "MERGE",
    "RENAME",
    "GRANT",
    "REVOKE",
]


def detect_dangerous_sql(sql: str):
    sql_upper = sql.upper()
    return [kw for kw in DANGEROUS_KEYWORDS if kw in sql_upper]


def is_read_only_sql(sql: str) -> bool:
    return sql.upper().strip().startswith(READ_ONLY_SQL_PREFIXES)


def sql_to_table_preview(sql: str):
    sql_upper = sql.upper()
    action = "UNKNOWN"
    table = "-"
    condition = "-"

    if sql_upper.startswith("DELETE"):
        action = "DELETE"
        match = re.search(r"FROM\s+(\w+)", sql_upper)
        if match:
            table = match.group(1)
        where_match = re.search(r"WHERE\s+(.+)", sql, re.IGNORECASE)
        if where_match:
            condition = where_match.group(1)

    return {
        "columns": ["Action", "Table", "Condition", "Impact"],
        "data": [[action, table, condition, "Removes record(s) permanently"]],
    }


def validate_sql_safety(sql: str) -> Tuple[bool, str]:
    """
    Fix for BUG 5: comments in LLM output used to cause an immediate hard
    security error with no retry. Now comments are stripped first (using
    the existing strip_sql_comments helper) so that an otherwise valid
    query is not rejected just because the LLM added a comment.
    """
    # Strip comments before validating — LLM sometimes adds them
    clean_sql = strip_sql_comments(sql).strip()

    clean_upper = clean_sql.upper().strip()
    if clean_upper.startswith(("DROP", "TRUNCATE")):
        return False, "DROP and TRUNCATE operations are not allowed"

    statements = clean_sql.split(";")
    if len([s for s in statements if s.strip()]) > 1:
        return False, "Multiple SQL statements are not allowed"

    return True, ""


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)


def get_user(identifier: str, db):
    return db.query(User).filter(
        or_(User.email == identifier, User.username == identifier)
    ).first()


def format_chat_history(chat_history: list) -> str:
    if not chat_history:
        return "No previous conversation history."
    recent_history = chat_history[-10:] if len(chat_history) > 10 else chat_history
    normalized_history = []

    if recent_history and isinstance(recent_history[0], dict) and "role" in recent_history[0]:
        pending_user = ""
        for item in recent_history:
            role = item.get("role")
            content = item.get("content", "")
            if role == "user":
                if pending_user:
                    normalized_history.append({"user": pending_user, "assistant": ""})
                pending_user = content
            elif role in {"assistant", "ai"}:
                normalized_history.append({"user": pending_user, "assistant": content})
                pending_user = ""
        if pending_user:
            normalized_history.append({"user": pending_user, "assistant": ""})
    else:
        normalized_history = recent_history[-5:]

    formatted = []
    for idx, item in enumerate(normalized_history[-5:], 1):
        user_msg = item.get('user', '')
        assistant_msg = item.get('assistant', '')
        sql_match = re.search(r'SQL: `([^`]+)`', assistant_msg)
        output_match = re.search(r'Output: ({.*})', assistant_msg, re.DOTALL)
        sql_query = sql_match.group(1) if sql_match else "N/A"
        if output_match:
            try:
                output_data = json.loads(output_match.group(1))
                if output_data.get('type') == 'select':
                    schema_info = f"Columns: {', '.join(output_data.get('columns', []))}, Rows: {output_data.get('row_count', 0)}"
                else:
                    schema_info = output_data.get('message', 'Operation completed')
            except Exception:
                schema_info = "Result schema unavailable"
        else:
            schema_info = "No output"
        formatted.append(
            f"\nPrevious Query #{idx}:\nUser Asked: \"{user_msg}\"\nSQL Generated: {sql_query}\nResult Schema: {schema_info}\n"
        )
    return "\n".join(formatted)


# ─────────────────────────────────────────────
# LLM CHAIN
# ─────────────────────────────────────────────
def is_association_table(columns: List[Dict[str, Any]], foreign_keys: List[Dict[str, Any]]) -> bool:
    if len(foreign_keys) < 2:
        return False

    column_names = {str(column.get("name", "")) for column in columns}
    ignored_columns = {"id", "created_at", "updated_at", "createdon", "updatedon", "deleted_at"}
    significant_columns = {name for name in column_names if name and name not in ignored_columns}
    fk_columns = {
        column_name
        for foreign_key in foreign_keys
        for column_name in (foreign_key.get("constrained_columns") or [])
    }
    return bool(significant_columns) and significant_columns.issubset(fk_columns)


def build_llm_schema_context(db) -> str:
    engine = db._engine
    inspector = sqlalchemy_inspect(engine)
    table_names = sorted(inspector.get_table_names())
    schema_lines: List[str] = []
    relationship_lines: List[str] = []

    for table_name in table_names:
        columns = inspector.get_columns(table_name)
        pk_constraint = inspector.get_pk_constraint(table_name) or {}
        primary_keys = set(pk_constraint.get("constrained_columns") or [])
        foreign_keys = inspector.get_foreign_keys(table_name) or []

        fk_map: Dict[str, List[str]] = {}
        for foreign_key in foreign_keys:
            constrained_columns = foreign_key.get("constrained_columns") or []
            referred_table = foreign_key.get("referred_table")
            referred_columns = foreign_key.get("referred_columns") or []

            for index, column_name in enumerate(constrained_columns):
                if referred_table and index < len(referred_columns):
                    fk_target = f"{referred_table}.{referred_columns[index]}"
                    fk_map.setdefault(column_name, []).append(fk_target)
                    relationship_lines.append(f"{table_name}.{column_name} -> {fk_target}")

        schema_lines.append(f"Table {table_name}:")
        for column in columns:
            column_name = str(column.get("name", ""))
            flags: List[str] = []
            if column_name in primary_keys:
                flags.append("PK")
            if not bool(column.get("nullable", True)):
                flags.append("NOT NULL")
            if column_name in fk_map:
                flags.append(f"FK -> {', '.join(fk_map[column_name])}")

            flags_text = f" [{', '.join(flags)}]" if flags else ""
            schema_lines.append(f"  - {column_name}: {column['type']}{flags_text}")

        if is_association_table(columns, foreign_keys):
            linked_tables = sorted({
                foreign_key.get("referred_table")
                for foreign_key in foreign_keys
                if foreign_key.get("referred_table")
            })
            if linked_tables:
                schema_lines.append(
                    f"  - Note: {table_name} is an association table linking {', '.join(linked_tables)}."
                )

        schema_lines.append("")

    summary_lines = [
        "Schema Summary:",
        f"- Dialect: {getattr(engine.dialect, 'name', 'sql')}",
        f"- Tables: {len(table_names)}",
        "",
    ]

    if relationship_lines:
        summary_lines.extend([
            "Known Relationships:",
            *[f"- {relationship}" for relationship in sorted(set(relationship_lines))],
            "",
        ])

    summary_lines.extend(schema_lines)
    logger.info(
        "Prepared schema context for LLM with %s tables and %s relationships.",
        len(table_names),
        len(set(relationship_lines)),
    )
    return "\n".join(summary_lines).strip()


def build_sql_prompt(
    db,
    question: str,
    chat_history: Optional[list] = None,
    failed_sql: Optional[str] = None,
    execution_error: Optional[str] = None,
) -> str:
    """
    Build the full LLM prompt.

    Fixes applied:
    - BUG 1: removed the broken .replace("{dialect}", ...) no-op.
    - BUG 3: dialect name and dialect-specific syntax rules are now
      explicitly injected into the prompt body so the LLM always
      knows which SQL dialect to produce.
    """
    history_context = format_chat_history(chat_history or [])
    raw_dialect = getattr(db._engine.dialect, "name", "sql").lower()

    dialect_key_map = {
        "mysql": "mysql",
        "mariadb": "mariadb",
        "postgresql": "postgresql",
        "oracle": "oracle",
        "mssql": "mssql",
        "sqlite": "sqlite",
        "ibm_db_sa": "db2",
    }
    dialect_key = dialect_key_map.get(raw_dialect, raw_dialect)
    dialect_rules = DIALECT_RULES.get(dialect_key, DEFAULT_DIALECT_RULES)
    dialect_display = raw_dialect.upper()

    system_prompt = SQL_SYSTEM_PROMPT.strip()
    schema_context = build_llm_schema_context(db)

    prompt_parts = [
        system_prompt,
        f"CURRENT TARGET DATABASE ENGINE: {dialect_display}",
        dialect_rules,
        "Database Schema:",
        schema_context,
        "Conversation History:",
        history_context,
        "Current User Question:",
        question.strip(),
    ]

    if failed_sql and execution_error:
        prompt_parts.extend([
            "Previous SQL Attempt (FAILED — do NOT repeat this):",
            failed_sql.strip(),
            "Execution Error:",
            execution_error.strip(),
            "Correction Rules:",
            f"- You must generate valid {dialect_display} SQL only.",
            "- Fix the SQL using only the exact tables and columns listed in the schema.",
            "- Re-check every JOIN key and every filtered column before answering.",
            "- If descriptive values are needed from a lookup table, join through the foreign keys shown in the schema.",
        ])
    else:
        prompt_parts.extend([
            "Generation Rules:",
            f"- Generate valid {dialect_display} SQL only.",
            "- Use the exact table names and column names from the schema.",
            "- Re-check every JOIN key and every filtered column before answering.",
            "- If descriptive values are needed from a lookup table, join through the foreign keys shown in the schema.",
        ])

    prompt_parts.extend([
        "Output Rules:",
        "- Return only one read-only SQL query.",
        "- No markdown fences.",
        "- No explanation.",
        "- No comments.",
        "- No trailing semicolon.",
        f"- The query MUST be valid {dialect_display} syntax.",
        "- The query MUST start with SELECT or WITH.",
    ])

    return "\n\n".join(prompt_parts)


def should_retry_sql_generation(error_message: str) -> bool:
    normalized_error = error_message.lower()
    retryable_markers = (
        "undefinedcolumn",
        "undefinedtable",
        "unknown column",
        "no such column",
        "no such table",
        "does not exist",
        "missing from-clause",
        "syntax error",
        "ambiguous column",
        "invalid reference",
    )
    return any(marker in normalized_error for marker in retryable_markers)


def format_query_execution_error(sql_query: str, error_message: str) -> Dict[str, Any]:
    if "only_full_group_by" in error_message or "1140" in error_message:
        helpful_msg = (
            "⚠️ GROUP BY Error: When using aggregate functions like COUNT, AVG, SUM, "
            "all non-aggregated columns must be included in the GROUP BY clause.\n\n"
            f"SQL attempted: {sql_query}\n\nTechnical error: {error_message}\n\n"
            "💡 Tip: Try rephrasing your question."
        )
        return {"type": "error", "message": helpful_msg}

    normalized_error = error_message.lower()
    if (
        "doesn't exist" in normalized_error
        or "does not exist" in normalized_error
        or "unknown column" in normalized_error
        or "no such column" in normalized_error
        or "no such table" in normalized_error
        or "undefinedcolumn" in normalized_error
        or "undefinedtable" in normalized_error
    ):
        helpful_msg = (
            "⚠️ Table/Column Not Found: The query references a table or column that doesn't exist.\n\n"
            f"SQL attempted: {sql_query}\n\nTechnical error: {error_message}\n\n"
            "💡 Tip: Ask me to show you the available tables and columns."
        )
        return {"type": "error", "message": helpful_msg}

    if "syntax error" in error_message.lower():
        helpful_msg = (
            f"⚠️ SQL Syntax Error: The generated query has a syntax problem.\n\n"
            f"SQL attempted: {sql_query}\n\nTechnical error: {error_message}\n\n"
            "💡 Tip: Try rephrasing your question."
        )
        return {"type": "error", "message": helpful_msg}

    return {
        "type": "error",
        "message": f"Query execution failed: {error_message}",
        "sql": sql_query,
    }


def execute_read_only_sql(sql_query: str, db, db_name: str) -> str:
    cached = query_cache.get(sql_query, db_name)
    if cached:
        return f"SQL: `{sql_query}`\nOutput: {json.dumps(cached)}"

    connection = None
    try:
        connection = db._engine.connect()
        result_proxy = None
        try:
            result_proxy = connection.execute(text(sql_query))
            columns = list(result_proxy.keys())
            rows = result_proxy.fetchmany(MAX_RESULT_ROWS)
            total_fetched = len(rows)

            data = [
                [str(cell) if cell is not None else '' for cell in row]
                for row in rows
            ]

            output_data = {
                "type": "select",
                "data": data,
                "columns": columns,
                "row_count": total_fetched,
                "limited": total_fetched == MAX_RESULT_ROWS,
            }
            query_cache.set(sql_query, db_name, output_data)
            return f"SQL: `{sql_query}`\nOutput: {json.dumps(output_data)}"
        finally:
            if result_proxy is not None:
                result_proxy.close()
    finally:
        if connection:
            try:
                connection.close()
            except Exception as exc:
                logger.warning("Error closing database connection: %s", exc)


def use_ollama() -> bool:
    return USE_OLLAMA.strip().lower() in {"1", "true", "yes", "on"}


def normalize_llm_text(content: Any) -> str:
    if isinstance(content, str):
        text_response = content.strip()
    elif isinstance(content, list):
        text_response = "".join(
            (part.get("text", "") if isinstance(part, dict) else getattr(part, "text", "")) or ""
            for part in content
        ).strip()
    else:
        text_response = str(content).strip()

    if not text_response:
        raise ValueError("LLM returned an empty response.")

    return text_response


def strip_sql_comments(sql: str) -> str:
    without_block_comments = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    without_dash_comments = re.sub(r"(?m)--[^\r\n]*$", "", without_block_comments)
    without_hash_comments = re.sub(r"(?m)#[^\r\n]*$", "", without_dash_comments)
    return without_hash_comments


def truncate_for_log(text: str, limit: int = 200) -> str:
    compact_text = re.sub(r"\s+", " ", text).strip()
    if len(compact_text) <= limit:
        return compact_text
    return compact_text[:limit].rstrip() + "..."


def prepare_prompt_for_ollama(prompt: str) -> str:
    sanitized_prompt = prompt.replace("\ufeff", "")
    sanitized_prompt = re.sub(r"[^\x09\x0A\x0D\x20-\x7E]", " ", sanitized_prompt)
    sanitized_prompt = re.sub(r"(?m)^\s*[-=*_]{8,}\s*$", "", sanitized_prompt)
    sanitized_prompt = re.sub(r"\n{3,}", "\n\n", sanitized_prompt).strip()

    ollama_suffix = (
        "\n\nStrict SQL-only response rule:\n"
        "Respond with exactly one read-only SQL query and nothing else.\n"
        "Do not repeat or explain the instructions.\n"
        "Do not describe allowed SQL keywords.\n"
        "Do not wrap the query in quotes.\n"
        "Do not add text like Here is the query or The generated result should be.\n"
        "Valid response example:\n"
        "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = DATABASE()"
    )
    return sanitized_prompt + ollama_suffix


def has_supported_read_only_start(text: str) -> bool:
    return re.match(
        r"^\s*(SELECT|WITH|SHOW|DESCRIBE|DESC|EXPLAIN|PRAGMA)\b(?=\s|\()",
        text,
        flags=re.IGNORECASE,
    ) is not None


def extract_sql_query(response_text: str) -> str:
    cleaned_response = re.sub(r"<think>.*?</think>", "", response_text, flags=re.IGNORECASE | re.DOTALL).strip()
    fenced_sql_match = re.search(r"```(?:sql)?\s*(.*?)```", cleaned_response, flags=re.IGNORECASE | re.DOTALL)
    if fenced_sql_match:
        cleaned_response = fenced_sql_match.group(1).strip()

    json_sql_match = re.search(
        r'"(?:sql|query)"\s*:\s*"((?:\\.|[^"\\])*)"',
        cleaned_response,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if json_sql_match:
        cleaned_response = bytes(json_sql_match.group(1), "utf-8").decode("unicode_escape").strip()
    else:
        quoted_sql_match = None
        for pattern in (
            r'"((?:\\.|[^"\\])*)"',
            r"`([^`]*)`",
        ):
            for match in re.finditer(pattern, cleaned_response, flags=re.DOTALL):
                candidate = match.group(1)
                if pattern.startswith(r'"'):
                    candidate = bytes(candidate, "utf-8").decode("unicode_escape")
                candidate = candidate.strip()
                if has_supported_read_only_start(candidate):
                    quoted_sql_match = candidate
                    break
            if quoted_sql_match:
                break

        if quoted_sql_match:
            cleaned_response = quoted_sql_match
        else:
            sql_start_match = re.search(
                r"\b(SELECT|WITH|SHOW|DESCRIBE|DESC|EXPLAIN|PRAGMA)\b(?=\s|\()",
                cleaned_response,
                flags=re.IGNORECASE,
            )
            if sql_start_match:
                cleaned_response = cleaned_response[sql_start_match.start():].strip()

    cleaned_response = re.sub(r"^\s*sql(?:\s+query)?\s*:\s*", "", cleaned_response, flags=re.IGNORECASE)
    cleaned_response = re.sub(r"^\s*(?:answer|output)\s*:\s*", "", cleaned_response, flags=re.IGNORECASE)
    cleaned_response = strip_sql_comments(cleaned_response).strip()

    statements = [statement.strip() for statement in cleaned_response.split(";") if statement.strip()]
    if len(statements) > 1:
        remaining_segments = statements[1:]
        contains_extra_sql = any(
            re.search(
                r"\b(SELECT|WITH|SHOW|DESCRIBE|DESC|EXPLAIN|PRAGMA|INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|MERGE|TRUNCATE|GRANT|REVOKE)\b",
                segment,
                flags=re.IGNORECASE,
            )
            for segment in remaining_segments
        )
        if contains_extra_sql:
            raise ValueError("LLM response contained multiple SQL statements.")
        logger.warning("Discarding trailing non-SQL text from LLM response.")
        cleaned_response = statements[0]
    elif statements:
        cleaned_response = statements[0]

    cleaned_response = cleaned_response.rstrip(";").strip()
    if not cleaned_response:
        raise ValueError("LLM response did not contain a valid SQL query.")

    if not has_supported_read_only_start(cleaned_response):
        raise ValueError("LLM response did not start with a supported read-only SQL statement.")

    return cleaned_response


def call_ollama(prompt: str) -> str:
    ollama_prompt = prepare_prompt_for_ollama(prompt)
    response = requests.post(
        OLLAMA_GENERATE_URL,
        json={
            "model": OLLAMA_MODEL,
            "prompt": ollama_prompt,
            "stream": False,
        },
        timeout=OLLAMA_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise ValueError("Ollama returned a non-JSON response payload.")
    return normalize_llm_text(payload.get("response"))


def is_ollama_alive() -> bool:
    if not use_ollama():
        return False

    try:
        response = requests.get(OLLAMA_HEALTH_URL, timeout=OLLAMA_HEALTH_TIMEOUT_SECONDS)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict):
            raise ValueError("Ollama health check returned an invalid payload.")

        models = payload.get("models", [])
        if isinstance(models, list) and models:
            model_names = [
                model.get("name", "") if isinstance(model, dict) else str(model)
                for model in models
            ]
            if not any(OLLAMA_MODEL in model_name for model_name in model_names):
                logger.warning("Ollama is reachable but model '%s' is not available.", OLLAMA_MODEL)
                return False

        return True
    except (requests.RequestException, ValueError) as exc:
        logger.warning("Ollama health check failed: %s", exc)
        return False


def call_groq(prompt: str) -> str:
    if not GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY is not configured.")

    llm = ChatGroq(
        api_key=SecretStr(str(GROQ_API_KEY)),
        model="llama-3.3-70b-versatile",
        temperature=0,
        max_tokens=6000,
        stop_sequences=None,
    )
    response = llm.invoke(prompt)
    return normalize_llm_text(getattr(response, "content", response))


def generate_llm_response(prompt: str, allow_ollama: bool = True) -> str:
    if allow_ollama and use_ollama():
        if is_ollama_alive():
            for attempt in range(OLLAMA_MAX_RETRIES + 1):
                try:
                    logger.info(
                        "Generating SQL with Ollama (attempt %s/%s).",
                        attempt + 1,
                        OLLAMA_MAX_RETRIES + 1,
                    )
                    return call_ollama(prompt)
                except (requests.RequestException, ValueError) as exc:
                    logger.warning(
                        "Ollama generation failed on attempt %s/%s: %s",
                        attempt + 1,
                        OLLAMA_MAX_RETRIES + 1,
                        exc,
                    )
            logger.warning("Ollama retries exhausted. Falling back to Groq.")
        else:
            logger.warning("Ollama is unavailable. Falling back to Groq.")
    elif use_ollama() and not allow_ollama:
        logger.info("Skipping Ollama and using Groq fallback for this generation attempt.")
    else:
        logger.info("USE_OLLAMA is disabled. Using Groq as the active LLM.")

    try:
        logger.info("Generating SQL with Groq fallback.")
        return call_groq(prompt)
    except Exception:
        logger.exception("Groq generation failed.")
        raise


def get_sql_chain(db, question: str, chat_history: Optional[list] = None) -> str:
    return build_sql_prompt(db, question, chat_history or [])


def get_response(question: str, db, db_name: str, chat_history: Optional[list] = None):
    """
    Generate SQL via LLM, execute it, return structured response string.

    Fix for BUG 2: when Ollama returns prose or invalid SQL and we fall
    back to Groq, the prompt is now rebuilt with the failed Ollama output
    attached as a "failed attempt". Previously the identical prompt was
    sent to Groq, which could produce the same broken output again.
    """
    sql_query = "N/A"
    prompt = get_sql_chain(db, question, chat_history or [])
    allow_ollama = True
    bad_ollama_text: Optional[str] = None

    try:
        for generation_attempt in range(2):
            response_text = ""
            try:
                response_text = generate_llm_response(prompt, allow_ollama=allow_ollama)
                sql_query = extract_sql_query(response_text)

            except ValueError as exc:
                if allow_ollama and use_ollama():
                    logger.info(
                        "Ollama response was not usable as SQL (%s). Rebuilding prompt for Groq fallback.",
                        exc,
                    )
                    bad_ollama_text = response_text
                    if bad_ollama_text:
                        logger.debug(
                            "Unusable Ollama response preview: %s",
                            truncate_for_log(bad_ollama_text),
                        )
                    allow_ollama = False
                    # Rebuild prompt with the bad Ollama output so Groq
                    # receives different input and knows to correct it.
                    groq_prompt = build_sql_prompt(
                        db,
                        question,
                        chat_history or [],
                        failed_sql=bad_ollama_text,
                        execution_error=(
                            "The previous attempt returned natural language or invalid SQL "
                            "instead of a valid query. Generate a correct SQL query now."
                        ),
                    )
                    response_text = generate_llm_response(groq_prompt, allow_ollama=False)
                    sql_query = extract_sql_query(response_text)
                else:
                    raise

            is_safe, error_msg = validate_sql_safety(sql_query)
            if not is_safe:
                if allow_ollama and use_ollama():
                    logger.warning(
                        "Ollama SQL failed safety validation (%s). Falling back to Groq. SQL: %s",
                        error_msg,
                        sql_query,
                    )
                    allow_ollama = False
                    groq_prompt = build_sql_prompt(
                        db,
                        question,
                        chat_history or [],
                        failed_sql=sql_query,
                        execution_error=f"Safety violation: {error_msg}",
                    )
                    response_text = generate_llm_response(groq_prompt, allow_ollama=False)
                    sql_query = extract_sql_query(response_text)
                    is_safe, error_msg = validate_sql_safety(sql_query)

                if not is_safe:
                    return json.dumps({
                        "type": "error",
                        "message": f"❌ Security Check Failed: {error_msg}",
                        "sql": sql_query,
                    })

            if not is_read_only_sql(sql_query):
                if allow_ollama and use_ollama():
                    logger.warning(
                        "Ollama SQL was not read-only. Falling back to Groq. SQL: %s",
                        sql_query,
                    )
                    allow_ollama = False
                    groq_prompt = build_sql_prompt(
                        db,
                        question,
                        chat_history or [],
                        failed_sql=sql_query,
                        execution_error=(
                            "The previous attempt generated a write/DDL statement. "
                            "Generate a read-only SELECT or WITH query."
                        ),
                    )
                    response_text = generate_llm_response(groq_prompt, allow_ollama=False)
                    sql_query = extract_sql_query(response_text)
                    is_safe, error_msg = validate_sql_safety(sql_query)
                    if not is_safe:
                        return json.dumps({
                            "type": "error",
                            "message": f"❌ Security Check Failed: {error_msg}",
                            "sql": sql_query,
                        })

                if not is_read_only_sql(sql_query):
                    return json.dumps({
                        "type": "error",
                        "sql": sql_query,
                        "message": (
                            "Write operations are temporarily disabled for security reasons. "
                            "Only read-only queries are allowed."
                        ),
                    })

            try:
                return execute_read_only_sql(sql_query, db, db_name)

            except Exception as select_error:
                error_message = str(select_error)
                if generation_attempt == 0 and should_retry_sql_generation(error_message):
                    logger.warning(
                        "Retrying SQL generation after execution failure. SQL: %s | Error: %s",
                        sql_query,
                        error_message,
                    )
                    prompt = build_sql_prompt(
                        db,
                        question,
                        chat_history or [],
                        failed_sql=sql_query,
                        execution_error=error_message,
                    )
                    if use_ollama():
                        allow_ollama = False
                    continue

                output_data = format_query_execution_error(sql_query, error_message)
                return f"SQL: `{sql_query}`\nOutput: {json.dumps(output_data)}"

    except Exception as exc:
        logger.exception("Error while generating or executing SQL: %s", exc)
        return json.dumps({
            "type": "error",
            "message": f"Error: {str(exc)}",
        })

    return json.dumps({
        "type": "error",
        "message": "Error: SQL generation finished without a result.",
    })


def run_chat_query(
    question: str,
    db_uri: str,
    db_name: str,
    db_type: str,
    chat_history: Optional[list] = None,
) -> str:
    if not source_supports_query(db_type):
        output_data = {
            "type": "error",
            "message": f"Natural-language querying is not enabled for {db_type} sources yet.",
        }
        return f"SQL: `N/A`\nOutput: {json.dumps(output_data)}"

    db = SQLDatabase(engine_cache.get(db_uri))
    return get_response(question=question, db=db, db_name=db_name, chat_history=chat_history or [])


# ─────────────────────────────────────────────
# AUTH ENDPOINTS
# ─────────────────────────────────────────────
@app.post("/api/send-otp")
@limiter.limit("5/minute")
async def send_otp_for_signup(request: Request, otp_request: OtpRequest, db: Session = Depends(get_db)):
    try:
        validate_email(otp_request.email)
    except EmailNotValidError as e:
        raise HTTPException(status_code=400, detail=f"Invalid email address: {str(e)}")

    otp = generate_otp()
    # OTP is NOT logged to stdout
    send_otp_email(otp_request.email, otp)
    otp_manager.store(otp_request.email, otp, db)
    return {"success": True, "message": "OTP has been sent to your email."}


@app.post("/api/signup", status_code=201)
async def signup_user(user: UserCreate, db: Session = Depends(get_db)):
    try:
        validate_email(user.email)
    except EmailNotValidError as e:
        raise HTTPException(status_code=400, detail=f"Invalid email address: {str(e)}")

    is_valid, error_msg = otp_manager.verify(user.email, user.otp, db)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    if get_user(user.email, db):
        raise HTTPException(status_code=400, detail="Email already registered")

    hashed_password = get_password_hash(user.password)
    db_user = User(
        email=user.email,
        firstName=user.firstName,
        lastName=user.lastName,
        gender=user.gender,
        username=user.username,
        hashed_password=hashed_password,
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    auth_token = create_auth_token(db_user.id, db)

    # Return user so frontend can auto-login without a second round-trip
    return {
        "success": True,
        "message": "User created successfully",
        "auth_token": auth_token,
        "user": {
            "id": db_user.id,
            "email": db_user.email,
            "firstName": db_user.firstName,
            "lastName": db_user.lastName,
            "username": db_user.username,
            "gender": db_user.gender,
        },
    }


@app.post("/api/login")
@limiter.limit("10/minute")
async def login_for_access_token(request: Request, form_data: UserLogin, db: Session = Depends(get_db)):
    user = get_user(form_data.identifier, db)
    if user is None or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    auth_token = create_auth_token(user.id, db)

    return {
        "success": True,
        "message": "Login successful",
        "auth_token": auth_token,
        "user": {
            "id": user.id,
            "email": user.email,
            "firstName": user.firstName,
            "lastName": user.lastName,
            "username": user.username,
            "gender": user.gender,
        },
    }


@app.get("/api/profile/{user_id}")
async def get_user_profile(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return {
        "success": True,
        "user": {
            "id": user.id,
            "email": user.email,
            "firstName": user.firstName,
            "lastName": user.lastName,
            "username": user.username,
            "gender": user.gender,
        },
    }


@app.post("/api/logout")
async def logout_user(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
):
    if authorization:
        try:
            revoke_auth_token(extract_bearer_token(authorization), db)
        except HTTPException:
            pass
    return {"success": True, "message": "Logged out successfully"}


# ─────────────────────────────────────────────
# DATABASE CONNECTION  (per-user session tokens)
# ─────────────────────────────────────────────
@app.post("/api/list-databases")
async def list_databases(
    config: ServerConnectionConfig,
    current_user: User = Depends(get_current_user),
):
    try:
        databases = list_available_databases_for_source(config)
        return {"success": True, "databases": databases, "total": len(databases)}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/create-database")
async def create_database(
    request: CreateDatabaseRequest,
    current_user: User = Depends(get_current_user),
):
    try:
        create_database_for_source(request)
        return {
            "success": True,
            "message": f"Database '{request.database_name}' created successfully",
            "database": request.database_name,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/connect")
async def connect_db(
    config: DBConfig,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Validates the connection and returns a per-user session token.
    The client MUST store this token and send it as X-DB-Session header
    on every subsequent API call that needs database access.
    """
    print(
        f"Connect request: type={config.type}, host={config.host}, port={config.port}, "
        f"database={config.database or config.path}"
    )
    try:
        source_type, db_uri, db_name = build_connection_artifacts(config)
        validate_source_connection(db_uri, source_type, db_name)
        token = db_session_store.create(current_user.id, db_uri, db_name, source_type, db)
        print("Database connection successful, session token issued")
        return {
            "success": True,
            "database": db_name,
            "type": source_type,
            "supports_query": source_supports_query(source_type),
            "session_token": token,
        }
    except Exception as e:
        print(f"Database connection failed: {str(e)}")
        return {"success": False, "error": str(e)}


@app.post("/api/connect-file")
async def connect_file_source(
    type: str = Form(...),
    database: str = Form(""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        source_type = normalize_source_type(type)
        if source_type not in FILE_SOURCE_TYPES:
            raise ValueError("This endpoint only supports CSV and Excel uploads.")

        db_uri = create_sql_snapshot_for_uploaded_file(source_type, file, current_user.id)
        db_name = database or file.filename or f"{source_type}_upload"
        validate_source_connection(db_uri, source_type, db_name)
        token = db_session_store.create(current_user.id, db_uri, db_name, source_type, db)
        return {
            "success": True,
            "database": db_name,
            "type": source_type,
            "supports_query": True,
            "session_token": token,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/disconnect")
async def disconnect_db(
    x_db_session: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if x_db_session:
        db_session_store.delete(x_db_session, db, current_user.id)
    query_cache.clear()
    return {"success": True, "message": "Database disconnected successfully"}


@app.get("/api/connection-status")
async def get_connection_status(
    x_db_session: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not x_db_session:
        return {
            "success": True,
            "connected": False,
            "database": None,
            "type": None,
            "supports_query": False,
        }
    session = db_session_store.get(x_db_session, db, current_user.id)
    if session is None:
        return {
            "success": True,
            "connected": False,
            "database": None,
            "type": None,
            "supports_query": False,
        }
    try:
        validate_source_connection(session["db_uri"], session["db_type"], session["db_name"])
        return {
            "success": True,
            "connected": True,
            "database": session["db_name"],
            "type": session["db_type"],
            "supports_query": source_supports_query(session["db_type"]),
        }
    except Exception:
        db_session_store.delete(x_db_session, db, current_user.id)
        return {
            "success": True,
            "connected": False,
            "database": None,
            "type": None,
            "supports_query": False,
        }


# ─────────────────────────────────────────────
# DATABASE SCHEMA / TABLE ENDPOINTS
# ─────────────────────────────────────────────
@app.get("/api/table-schema/{table_name}")
async def get_table_schema(
    table_name: str,
    db_session: Dict[str, Any] = Depends(get_db_session),
):
    try:
        source_type = normalize_source_type(db_session["db_type"])
        if source_type in SQL_SOURCE_TYPES:
            columns = get_sql_table_schema(db_session["db_uri"], table_name)
        elif source_type == "mongodb":
            columns = get_mongodb_table_schema(db_session["db_uri"], db_session["db_name"], table_name)
        elif source_type == "redis":
            columns = get_redis_table_schema()
        else:
            raise ValueError(f"Unsupported source type: {source_type}")
        return {"success": True, "table_name": table_name, "columns": columns}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/database-tables")
async def get_database_tables(db_session: Dict[str, Any] = Depends(get_db_session)):
    try:
        source_type = normalize_source_type(db_session["db_type"])
        if source_type in SQL_SOURCE_TYPES:
            tables_info = get_sql_tables_info(db_session["db_uri"])
        elif source_type == "mongodb":
            tables_info = get_mongodb_tables_info(db_session["db_uri"], db_session["db_name"])
        elif source_type == "redis":
            tables_info = get_redis_tables_info(db_session["db_uri"])
        else:
            raise ValueError(f"Unsupported source type: {source_type}")
        return {"success": True, "tables": tables_info, "total": len(tables_info)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
# CHAT ENDPOINT
# ─────────────────────────────────────────────
@app.post("/api/chat")
@limiter.limit("60/minute")
async def chat_endpoint(
    request: Request,
    chat_request: ChatRequest,
    db_session: Dict[str, Any] = Depends(get_db_session),
):
    try:
        db_uri = db_session["db_uri"]
        db_name = db_session["db_name"]
        response = await run_in_threadpool(
            run_chat_query,
            chat_request.question,
            db_uri,
            db_name,
            db_session["db_type"],
            chat_request.chat_history,
        )
        return {"success": True, "response": response}
    except HTTPException as e:
        raise e
    except Exception as exc:
        logger.exception("Chat endpoint error: %s", exc)
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(exc)}")


# ─────────────────────────────────────────────
# CONFIRM-SQL  — DISABLED until properly secured
# ─────────────────────────────────────────────
@app.post("/api/confirm-sql")
async def confirm_sql_action():
    """
    Write-SQL execution is temporarily disabled pending full
    server-side pending-action implementation with audit logging.
    """
    raise HTTPException(
        status_code=503,
        detail="Write operations are temporarily disabled. Contact support for updates.",
    )


# ─────────────────────────────────────────────
# CHAT SESSIONS
# NOTE: delete-all MUST be declared before /{session_id}
# ─────────────────────────────────────────────
@app.get("/api/chat-sessions")
async def get_chat_sessions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sessions = db.query(ChatSession).filter(ChatSession.user_id == current_user.id).all()
    result = []
    for session in sessions:
        result.append({
            "id": session.id,
            "title": session.title,
            "messages": json.loads(str(session.messages)),
            "timestamp": datetime.utcnow().isoformat(),
        })
    return result


@app.post("/api/chat-sessions")
async def create_chat_session(
    session: ChatSessionCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(session.user_id, current_user)
    try:
        new_session = ChatSession(
            user_id=current_user.id,
            title=session.title or "Untitled Chat",
            messages=json.dumps(session.messages),
        )
        db.add(new_session)
        db.commit()
        db.refresh(new_session)
        return {
            "id": new_session.id,
            "title": new_session.title,
            "messages": json.loads(str(new_session.messages)),
            "timestamp": datetime.utcnow().isoformat(),
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create chat session: {str(e)}")


@app.put("/api/chat-sessions/{session_id}")
async def update_chat_session(
    session_id: int,
    session: ChatSessionUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(session.user_id, current_user)
    try:
        existing_session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
        if not existing_session:
            raise HTTPException(status_code=404, detail="Chat session not found")
        if cast(int, existing_session.user_id) != current_user.id:
            raise HTTPException(status_code=403, detail="Unauthorized to update this session")
        if session.title is not None:
            existing_session.title = session.title
        if session.messages is not None:
            setattr(existing_session, 'messages', json.dumps(session.messages))
        db.commit()
        return {
            "id": existing_session.id,
            "title": existing_session.title,
            "messages": json.loads(str(existing_session.messages)),
            "timestamp": datetime.utcnow().isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update chat session: {str(e)}")


# IMPORTANT: delete-all must come BEFORE the dynamic {session_id} route
@app.delete("/api/chat-sessions/delete-all")
async def delete_all_chat_sessions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        sessions = db.query(ChatSession).filter(ChatSession.user_id == current_user.id).all()
        if not sessions:
            return {"success": True, "message": "No chat sessions to delete"}
        for session in sessions:
            db.delete(session)
        db.commit()
        return {"success": True, "message": f"Deleted {len(sessions)} chat sessions"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete chat sessions: {str(e)}")


@app.delete("/api/chat-sessions/{session_id}")
async def delete_chat_session(
    session_id: int = Path(..., description="The ID of the chat session to delete"),
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    try:
        chat_session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
        if not chat_session:
            raise HTTPException(status_code=404, detail="Chat session not found")
        if cast(int, chat_session.user_id) != current_user.id:
            raise HTTPException(status_code=403, detail="Unauthorized to delete this session")
        db.delete(chat_session)
        db.commit()
        return {"success": True, "message": "Chat session deleted"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete chat session: {str(e)}")


# ─────────────────────────────────────────────
# FAVORITES
# ─────────────────────────────────────────────
@app.get("/api/favorites/{user_id}")
async def get_favorites(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    favorites = db.query(FavoriteQuery).filter(
        FavoriteQuery.user_id == user_id
    ).order_by(FavoriteQuery.created_at.desc()).all()
    return [{
        "id": fav.id,
        "question": fav.question,
        "sql_query": fav.sql_query,
        "tags": fav.tags,
        "description": fav.description,
        "created_at": fav.created_at.isoformat(),
    } for fav in favorites]


@app.get("/api/favorites/{user_id}/check")
async def check_favorite(
    user_id: int,
    sql: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    favorite = db.query(FavoriteQuery).filter(
        FavoriteQuery.user_id == user_id,
        FavoriteQuery.sql_query == sql,
    ).first()
    return {
        "is_favorite": favorite is not None,
        "favorite_id": favorite.id if favorite else None,
    }


@app.post("/api/favorites")
async def add_favorite(
    favorite: FavoriteQueryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(favorite.user_id, current_user)
    existing = db.query(FavoriteQuery).filter(
        FavoriteQuery.user_id == current_user.id,
        FavoriteQuery.sql_query == favorite.sql_query,
    ).first()
    if existing:
        return {"success": False, "message": "Query already in favorites"}
    new_favorite = FavoriteQuery(
        user_id=current_user.id,
        question=favorite.question,
        sql_query=favorite.sql_query,
        tags=favorite.tags,
        description=favorite.description,
    )
    db.add(new_favorite)
    db.commit()
    db.refresh(new_favorite)
    return {"success": True, "message": "Added to favorites", "id": new_favorite.id}


@app.delete("/api/favorites/{favorite_id}")
async def remove_favorite(
    favorite_id: int,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    favorite = db.query(FavoriteQuery).filter(
        FavoriteQuery.id == favorite_id,
        FavoriteQuery.user_id == current_user.id,
    ).first()
    if not favorite:
        raise HTTPException(status_code=404, detail="Favorite not found")
    db.delete(favorite)
    db.commit()
    return {"success": True, "message": "Removed from favorites"}


# ─────────────────────────────────────────────
# USER SETTINGS
# ─────────────────────────────────────────────
@app.get("/api/settings/{user_id}")
async def get_user_settings(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    settings = db.query(UserSettings).filter(UserSettings.user_id == user_id).first()
    if not settings:
        settings = UserSettings(user_id=user_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return {
        "theme": settings.theme,
        "language": settings.language,
        "results_per_page": settings.results_per_page,
        "auto_save_sessions": settings.auto_save_sessions,
        "sql_syntax_highlighting": settings.sql_syntax_highlighting,
        "notification_preferences": settings.notification_preferences or {},
    }


@app.put("/api/settings/{user_id}")
async def update_user_settings(
    user_id: int,
    settings_update: UserSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    settings = db.query(UserSettings).filter(UserSettings.user_id == user_id).first()
    if not settings:
        settings = UserSettings(user_id=user_id)
        db.add(settings)
    update_data = settings_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(settings, field, value)
    setattr(settings, 'updated_at', datetime.utcnow())
    db.commit()
    db.refresh(settings)
    return {"success": True, "message": "Settings updated successfully"}


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────
@app.post("/api/export")
async def export_results(request: ExportRequest):
    try:
        if request.format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(request.columns)
            writer.writerows(request.data)
            return {
                "success": True,
                "format": "csv",
                "data": output.getvalue(),
                "filename": f"query_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            }
        elif request.format == "json":
            results = []
            for row in request.data:
                row_dict = {}
                for i, col in enumerate(request.columns):
                    row_dict[col] = row[i] if i < len(row) else None
                results.append(row_dict)
            return {
                "success": True,
                "format": "json",
                "data": json.dumps(results, indent=2),
                "filename": f"query_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            }
        else:
            raise HTTPException(status_code=400, detail="Unsupported export format. Use 'csv' or 'json'")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ─────────────────────────────────────────────
# SEARCH TABLES
# ─────────────────────────────────────────────
@app.get("/api/search/tables")
async def search_tables(query: str, db_session: Dict[str, Any] = Depends(get_db_session)):
    try:
        db = SQLDatabase(engine_cache.get(db_session["db_uri"]))
        schema_info = db.get_table_info()
        query_lower = query.lower()
        results = {"tables": [], "columns": []}
        lines = schema_info.split('\n')
        current_table = None
        for line in lines:
            if 'CREATE TABLE' in line:
                table_name = line.split('`')[1] if '`' in line else None
                current_table = table_name
                if table_name and query_lower in table_name.lower():
                    results["tables"].append(table_name)
            elif current_table and '`' in line:
                col_name = line.split('`')[1]
                if query_lower in col_name.lower():
                    results["columns"].append({"table": current_table, "column": col_name})
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


# ─────────────────────────────────────────────
# QUERY HISTORY
# ─────────────────────────────────────────────
@app.post("/api/history/track")
async def track_query_execution(
    data: QueryHistoryTrackRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(data.user_id, current_user)
    history = QueryHistory(
        user_id=current_user.id,
        session_id=data.session_id,
        question=data.question,
        sql_query=data.sql_query,
        success=data.success,
        execution_time_ms=data.execution_time_ms,
        row_count=data.row_count,
    )
    db.add(history)
    db.commit()
    return {"success": True, "message": "Query tracked"}


@app.get("/api/history/{user_id}")
async def get_query_history(
    user_id: int,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    history = db.query(QueryHistory).filter(
        QueryHistory.user_id == user_id
    ).order_by(QueryHistory.created_at.desc()).limit(limit).all()
    return [{
        "id": h.id,
        "question": h.question,
        "sql_query": h.sql_query,
        "success": h.success,
        "execution_time_ms": h.execution_time_ms,
        "row_count": h.row_count,
        "created_at": h.created_at.isoformat(),
    } for h in history]


@app.get("/api/history/{user_id}/stats")
async def get_query_stats(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    total_queries = db.query(QueryHistory).filter(QueryHistory.user_id == user_id).count()
    successful_queries = db.query(QueryHistory).filter(
        QueryHistory.user_id == user_id,
        QueryHistory.success.is_(True),
    ).count()
    return {
        "total_queries": total_queries,
        "successful_queries": successful_queries,
        "success_rate": (successful_queries / total_queries * 100) if total_queries > 0 else 0,
    }


# ─────────────────────────────────────────────
# CUSTOM DASHBOARDS
# ─────────────────────────────────────────────
def serialize_dashboard(dashboard: UserDashboard) -> Dict[str, Any]:
    return {
        "id": dashboard.id,
        "dashboard_id": dashboard.dashboard_id,
        "user_id": dashboard.user_id,
        "name": dashboard.name,
        "description": dashboard.description,
        "charts": json.loads(dashboard.charts_data),
        "created_at": dashboard.created_at.isoformat() if dashboard.created_at else None,
        "updated_at": dashboard.updated_at.isoformat() if dashboard.updated_at else None,
    }


@app.get("/api/custom-dashboards/{user_id}")
async def get_user_dashboards(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    dashboards = db.query(UserDashboard).filter(UserDashboard.user_id == user_id).all()
    return [serialize_dashboard(dashboard) for dashboard in dashboards]


@app.post("/api/custom-dashboards")
async def create_dashboard(
    dashboard: DashboardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(dashboard.user_id, current_user)
    existing = db.query(UserDashboard).filter(
        UserDashboard.dashboard_id == dashboard.dashboard_id,
        UserDashboard.user_id == current_user.id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Dashboard ID already exists")
    new_dashboard = UserDashboard(
        user_id=current_user.id,
        dashboard_id=dashboard.dashboard_id,
        name=dashboard.name,
        description=dashboard.description or '',
        charts_data=json.dumps(dashboard.charts),
    )
    db.add(new_dashboard)
    db.commit()
    db.refresh(new_dashboard)
    return serialize_dashboard(new_dashboard)


@app.put("/api/custom-dashboards/{dashboard_id}")
async def update_dashboard(
    dashboard_id: str,
    updates: DashboardUpdate,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    dashboard = db.query(UserDashboard).filter(
        UserDashboard.dashboard_id == dashboard_id,
        UserDashboard.user_id == current_user.id,
    ).first()
    if not dashboard:
        raise HTTPException(status_code=404, detail="Dashboard not found")
    if updates.name is not None:
        dashboard.name = updates.name
    if updates.description is not None:
        dashboard.description = updates.description
    if updates.charts is not None:
        dashboard.charts_data = json.dumps(updates.charts)
    dashboard.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(dashboard)
    return serialize_dashboard(dashboard)


@app.delete("/api/custom-dashboards/{dashboard_id}")
async def delete_dashboard(
    dashboard_id: str,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(user_id, current_user)
    dashboard = db.query(UserDashboard).filter(
        UserDashboard.dashboard_id == dashboard_id,
        UserDashboard.user_id == current_user.id,
    ).first()
    if not dashboard:
        raise HTTPException(status_code=404, detail="Dashboard not found")
    db.delete(dashboard)
    db.commit()
    return {"success": True, "message": "Dashboard deleted"}


@app.post("/api/custom-dashboards/migrate-from-localstorage")
async def migrate_dashboards(
    migration: DashboardMigrate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    assert_user_access(migration.user_id, current_user)
    try:
        migrated_count = 0
        skipped_count = 0
        for dashboard_data in migration.dashboards_data:
            existing = db.query(UserDashboard).filter(
                UserDashboard.dashboard_id == dashboard_data.get('id'),
                UserDashboard.user_id == current_user.id,
            ).first()
            if existing:
                skipped_count += 1
                continue
            new_dashboard = UserDashboard(
                user_id=current_user.id,
                dashboard_id=dashboard_data.get('id'),
                name=dashboard_data.get('name', 'Untitled Dashboard'),
                description=dashboard_data.get('description', ''),
                charts_data=json.dumps(dashboard_data.get('charts', [])),
                created_at=datetime.fromisoformat(
                    dashboard_data.get('createdAt', datetime.utcnow().isoformat()).replace('Z', '+00:00')
                ),
                updated_at=datetime.fromisoformat(
                    dashboard_data.get('updatedAt', datetime.utcnow().isoformat()).replace('Z', '+00:00')
                ),
            )
            db.add(new_dashboard)
            migrated_count += 1
        db.commit()
        message = f"Migrated {migrated_count} dashboards, skipped {skipped_count} existing"
        return {"success": True, "message": message, "migrated": migrated_count, "skipped": skipped_count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.on_event("shutdown")
async def shutdown_event():
    engine_cache.dispose_all()
    print("Application shutdown - resources cleaned up")
